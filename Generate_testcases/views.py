from django.shortcuts import render,HttpResponse

# Create your views here.

# app/views.py
from django.shortcuts import get_object_or_404, redirect, render
from django.db import transaction
from django.urls import reverse
from django.contrib import messages
from django.http import JsonResponse
from django.views.decorators.http import require_http_methods
from django.conf import settings

from .models import (
    FeatureLevel1, FeatureLevel2, TestCaseSeed, 
    GenerationSession, GenerationItem, GenerationSeedConfig,
    SavedCaseItem
)
from .forms import (
    FeatureLevel1Form, FeatureLevel2Form, SeedSelectionForm,
    GenerationSessionForm, GenerationItemFormSet, SaveCaseSetForm, TestCaseSeedForm
)
from .llm_client import generate_cases_for_seed, LLMError

# views.py 末尾追加
# from django.http import JsonResponse
# from django.views.decorators.http import require_http_methods
# from django.db import transaction

from openpyxl import load_workbook

from .models import FeatureLevel1, FeatureLevel2, TestCaseSeed

def testcase_workspace(request):
    """
    测试用例工作台：三栏布局
    - 左侧：一级功能列表
    - 中间：二级功能列表
    - 右侧：种子测试用例列表和生成
    """
    # 获取所有一级功能
    level1_list = FeatureLevel1.objects.all().order_by("name")
    
    context = {
        "level1_list": level1_list,
    }
    return render(request, "testcase_workspace.html", context)


def level2_list(request):
    """
    二级功能列表页
    """
    qs = FeatureLevel2.objects.select_related("level1").order_by("level1__name", "name")
    return render(request, "level2_list.html", {"level2_list": qs})


def create_or_select_scenario(request):
    """
    创建或选择场景（一级功能）和二级功能
    支持：
    1. 创建新的一级功能和二级功能
    2. 选择已存在的一级功能和二级功能
    3. 选择多个种子并指定每个种子的生成数量
    """
    level1_form = FeatureLevel1Form()
    level2_form = None
    seed_form = None
    level1 = None
    level2 = None
    
    if request.method == "POST":
        action = request.POST.get("action")
        
        # 步骤1：处理一级功能
        if action == "step1_level1":
            level1_form = FeatureLevel1Form(request.POST)
            if level1_form.is_valid():
                use_existing = level1_form.cleaned_data.get("use_existing", False)
                existing_level1_obj = level1_form.cleaned_data.get("existing_level1")
                
                if use_existing and existing_level1_obj:
                    # existing_level1_obj 是一个 FeatureLevel1 对象
                    level1 = existing_level1_obj
                    messages.info(request, f"已选择一级功能：{level1.name}")
                else:
                    # 创建新的一级功能，先检查是否已存在同名功能
                    level1_name = level1_form.cleaned_data.get("name")
                    existing_level1 = FeatureLevel1.objects.filter(name=level1_name).first()
                    
                    if existing_level1:
                        # 如果已存在，直接使用现有的
                        level1 = existing_level1
                        messages.warning(request, f"一级功能「{level1.name}」已存在，已自动选择该功能")
                    else:
                        # 不存在则创建新的
                        level1 = level1_form.save()
                        messages.success(request, f"已创建一级功能：{level1.name}")
                
                # 初始化二级功能表单
                level2_form = FeatureLevel2Form(level1=level1)
        
        # 步骤2：处理二级功能
        elif action == "step2_level2":
            level1_id = request.POST.get("level1_id")
            level1 = get_object_or_404(FeatureLevel1, id=level1_id)
            level2_form = FeatureLevel2Form(request.POST, level1=level1)
            
            if level2_form.is_valid():
                use_existing = level2_form.cleaned_data.get("use_existing", False)
                existing_level2_obj = level2_form.cleaned_data.get("existing_level2")
                
                if use_existing and existing_level2_obj:
                    # existing_level2_obj 是一个 FeatureLevel2 对象
                    level2 = existing_level2_obj
                    messages.info(request, f"已选择二级功能：{level2.name}")
                else:
                    # 创建新的二级功能，先检查是否已存在
                    level2_name = level2_form.cleaned_data.get("name")
                    existing_level2 = FeatureLevel2.objects.filter(level1=level1, name=level2_name).first()
                    
                    if existing_level2:
                        # 如果已存在，直接使用现有的
                        level2 = existing_level2
                        messages.warning(request, f"二级功能「{level2.name}」已存在，已自动选择该功能")
                    else:
                        # 不存在则创建新的
                        level2 = level2_form.save(commit=False)
                        level2.level1 = level1
                        level2.save()
                        messages.success(request, f"已创建二级功能：{level2.name}")
                
                # 初始化种子选择表单
                seed_form = SeedSelectionForm(level2=level2)
                session_form = GenerationSessionForm(initial={"temperature": 0.7, "top_p": 1.0})
        
        # 步骤3：生成测试用例
        elif action == "step3_generate":
            level2_id = request.POST.get("level2_id")
            level2 = get_object_or_404(FeatureLevel2.objects.select_related("level1"), id=level2_id)
            level1 = level2.level1
            
            seed_form = SeedSelectionForm(request.POST, level2=level2)
            session_form = GenerationSessionForm(request.POST)
            
            if seed_form.is_valid() and session_form.is_valid():
                selected_seeds = seed_form.get_selected_seeds()
                
                if not selected_seeds:
                    messages.error(request, "请至少选择一个种子样例")
                    seed_form = SeedSelectionForm(level2=level2)
                else:
                    with transaction.atomic():
                        # # 创建生成会话
                        # session = session_form.save(commit=False)
                        # session.level2 = level2
                        # session.model_name = "mock-model"
                        # session.status = "done"
                        # session.created_by = request.user if request.user.is_authenticated else None
                        # session.save()
                        #
                        # # 为每个选中的种子创建配置并生成用例
                        # idx = 0
                        # for seed, n in selected_seeds:
                        #     # 创建种子配置
                        #     GenerationSeedConfig.objects.create(
                        #         session=session,
                        #         seed=seed,
                        #         n=n
                        #     )
                        #
                        #     # 生成用例（mock）
                        #     base = seed.text[:50] + "…" if len(seed.text) > 50 else seed.text
                        #     effective_prompt = session.effective_prompt
                        #     for i in range(n):
                        #         GenerationItem.objects.create(
                        #             session=session,
                        #             idx=idx,
                        #             raw_text=f"[{level2.name}] 泛化用例 {idx+1} | seed={base} | prompt={effective_prompt[:30] if effective_prompt else 'default'}",
                        #         )
                        #         idx += 1
                        # 创建生成会话
                        session = session_form.save(commit=False)
                        session.level2 = level2

                        # 你要求“只有一个prompt”：这里以 session.prompt 为准（来自前端提交）
                        # session_form 已包含 prompt 字段 :contentReference[oaicite:10]{index=10}

                        session.model_name = "your-llm-model"  # 这里写你实际模型名/渠道名
                        session.status = "draft"
                        session.created_by = request.user if request.user.is_authenticated else None
                        session.save()

                        idx = 0
                        try:
                            for seed, n in selected_seeds:
                                GenerationSeedConfig.objects.create(session=session, seed=seed, n=n)

                                # === 真实大模型调用 ===
                                prompt = (session.prompt or "").strip()  # 你定义的“唯一prompt”
                                outputs = generate_cases_by_llm(
                                    level1_name=level1.name,
                                    level2_name=level2.name,
                                    prompt=prompt,
                                    seed_text=seed.text,
                                    n=n,
                                    temperature=session.temperature,
                                    top_p=session.top_p,
                                )

                                # outputs 必须是 list[str]，长度最好==n（不足就按实际写入）
                                for text in outputs:
                                    text = (text or "").strip()
                                    if not text:
                                        continue
                                    GenerationItem.objects.create(
                                        session=session,
                                        seed=seed,  # ✅ 关键：补上seed关联，否则结果页按种子分组会丢 :contentReference[oaicite:11]{index=11}
                                        idx=idx,
                                        raw_text=text,  # ✅ 真实模型输出直接写这里
                                    )
                                    idx += 1

                            session.status = "done"
                            session.save(update_fields=["status"])

                        except Exception as e:
                            # 失败要落库，方便前端提示/排查
                            session.status = "failed"
                            session.save(update_fields=["status"])
                            raise

                    messages.success(request, f"生成完成！共生成 {idx} 条用例（mock）")
                    return redirect(reverse("Generate_testcases:level2_detail", args=[level2.id]))
            else:
                # 表单验证失败，重新显示种子选择表单
                seed_form = SeedSelectionForm(level2=level2)
    
    # GET请求或需要重新显示表单
    # 准备种子列表（手动处理以便模板更简单）
    seed_list = []
    if level2 and seed_form:
        seeds = TestCaseSeed.objects.filter(level2=level2).order_by("-created_at")
        for seed in seeds:
            seed_list.append({
                'id': seed.id,
                'text': seed.text,
                'field_name': f'seed_{seed.id}',
                'n_field_name': f'seed_{seed.id}_n'
            })
    
    # 确保session_form存在
    if 'session_form' not in locals():
        session_form = GenerationSessionForm(initial={"temperature": 0.7, "top_p": 1.0}) if level2 else None
    
    context = {
        "level1_form": level1_form,
        "level2_form": level2_form,
        "seed_form": seed_form,
        "seed_list": seed_list,
        "session_form": session_form,
        "level1": level1,
        "level2": level2,
    }
    return render(request, "create_or_select_scenario.html", context)


@require_http_methods(["GET"])
def get_level2_list(request):
    """AJAX接口：根据一级功能ID获取二级功能列表"""
    level1_id = request.GET.get("level1_id")
    if not level1_id:
        return JsonResponse({"error": "缺少level1_id参数"}, status=400)

    try:
        level1 = FeatureLevel1.objects.get(id=level1_id)
        level2_list = FeatureLevel2.objects.filter(level1=level1).order_by("name")
        data = [{"id": l2.id, "name": l2.name, "prompt": l2.prompt or ""} for l2 in level2_list]
        return JsonResponse({"level2_list": data})
    except FeatureLevel1.DoesNotExist:
        return JsonResponse({"error": "一级功能不存在"}, status=404)


@require_http_methods(["GET"])
def get_seed_list(request):
    """AJAX接口：根据二级功能ID获取种子测试用例列表"""
    level2_id = request.GET.get("level2_id")
    if not level2_id:
        return JsonResponse({"error": "缺少level2_id参数"}, status=400)
    
    try:
        level2 = FeatureLevel2.objects.get(id=level2_id)
        seeds = TestCaseSeed.objects.filter(level2=level2).order_by("-created_at")
        data = [{
            "id": seed.id,
            "text": seed.text,
            "created_at": seed.created_at.strftime("%Y-%m-%d %H:%M")
        } for seed in seeds]
        return JsonResponse({"seed_list": data, "prompt": level2.prompt or ""})
    except FeatureLevel2.DoesNotExist:
        return JsonResponse({"error": "二级功能不存在"}, status=404)


@require_http_methods(["POST"])
def add_level1(request):
    """AJAX接口：添加一级功能"""
    name = request.POST.get("name", "").strip()
    code = request.POST.get("code", "").strip()
    
    if not name:
        return JsonResponse({"error": "名称不能为空"}, status=400)
    
    # 检查是否已存在
    existing = FeatureLevel1.objects.filter(name=name).first()
    if existing:
        return JsonResponse({"error": f"一级功能「{name}」已存在"}, status=400)
    
    level1 = FeatureLevel1.objects.create(name=name, code=code)
    return JsonResponse({"id": level1.id, "name": level1.name, "message": "添加成功"})


@require_http_methods(["POST"])
def add_level2(request):
    """AJAX接口：添加二级功能"""
    level1_id = request.POST.get("level1_id")
    name = request.POST.get("name", "").strip()
    code = request.POST.get("code", "").strip()
    prompt = request.POST.get("prompt", "").strip()
    
    if not level1_id:
        return JsonResponse({"error": "缺少一级功能ID"}, status=400)
    if not name:
        return JsonResponse({"error": "名称不能为空"}, status=400)
    
    try:
        level1 = FeatureLevel1.objects.get(id=level1_id)
    except FeatureLevel1.DoesNotExist:
        return JsonResponse({"error": "一级功能不存在"}, status=404)
    
    # 检查是否已存在
    existing = FeatureLevel2.objects.filter(level1=level1, name=name).first()
    if existing:
        return JsonResponse({"error": f"二级功能「{name}」已存在"}, status=400)
    
    level2 = FeatureLevel2.objects.create(level1=level1, name=name, code=code, prompt=prompt)
    return JsonResponse({"id": level2.id, "name": level2.name, "prompt": level2.prompt or "", "message": "添加成功"})


@require_http_methods(["POST"])
def add_seed(request):
    """AJAX接口：添加种子测试用例"""
    level2_id = request.POST.get("level2_id")
    text = request.POST.get("text", "").strip()
    
    if not level2_id:
        return JsonResponse({"error": "缺少二级功能ID"}, status=400)
    if not text:
        return JsonResponse({"error": "测试用例内容不能为空"}, status=400)
    
    try:
        level2 = FeatureLevel2.objects.get(id=level2_id)
    except FeatureLevel2.DoesNotExist:
        return JsonResponse({"error": "二级功能不存在"}, status=404)
    
    seed = TestCaseSeed.objects.create(
        level2=level2,
        text=text,
        source="manual",
        created_by=request.user if request.user.is_authenticated else None
    )
    
    return JsonResponse({
        "id": seed.id,
        "text": seed.text,
        "created_at": seed.created_at.strftime("%Y-%m-%d %H:%M"),
        "message": "添加成功"
    })


# @require_http_methods(["POST"])
# def workspace_generate(request):
#     """工作台生成测试用例"""
#     level2_id = request.POST.get("level2_id")
#     seed_configs = request.POST.get("seed_configs")  # JSON格式: [{"seed_id": 1, "n": 5}, ...]
#     prompt = request.POST.get("prompt", "").strip()
#     temperature = float(request.POST.get("temperature", 0.7))
#     top_p = float(request.POST.get("top_p", 1.0))
#
#     if not level2_id:
#         return JsonResponse({"error": "缺少二级功能ID"}, status=400)
#     if not seed_configs:
#         return JsonResponse({"error": "请至少选择一个种子测试用例"}, status=400)
#
#     try:
#         import json
#         seed_configs = json.loads(seed_configs)
#
#         print("### workspace_generate: seed_configs_len =", len(seed_configs))
#         print("### workspace_generate: seed_configs =", seed_configs[:5], "...")
#
#     except:
#         return JsonResponse({"error": "种子配置格式错误"}, status=400)
#
#     try:
#         level2 = FeatureLevel2.objects.get(id=level2_id)
#     except FeatureLevel2.DoesNotExist:
#         return JsonResponse({"error": "二级功能不存在"}, status=404)
#
#     with transaction.atomic():
#         # 创建生成会话
#         session = GenerationSession.objects.create(
#             level2=level2,
#             prompt=prompt or None,
#             model_name="mock-model",
#             temperature=temperature,
#             top_p=top_p,
#             status="done",
#             created_by=request.user if request.user.is_authenticated else None
#         )
#
#         # 为每个种子创建配置并生成用例
#         idx = 0
#         for config in seed_configs:
#             seed_id = config.get("seed_id")
#             n = config.get("n", 5)
#
#             try:
#                 seed = TestCaseSeed.objects.get(id=seed_id, level2=level2)
#             except TestCaseSeed.DoesNotExist:
#                 continue
#
#             # 创建种子配置
#             # GenerationSeedConfig.objects.create(
#             #     session=session,
#             #     seed=seed,
#             #     n=n
#             # )
#             GenerationSeedConfig.objects.update_or_create(
#                 session=session,
#                 seed=seed,
#                 defaults={"n": n},
#             )
#
#             # # 生成用例（mock）
#             # base = seed.text[:50] + "…" if len(seed.text) > 50 else seed.text
#             # effective_prompt = session.effective_prompt
#             # for i in range(n):
#             #     GenerationItem.objects.create(
#             #         session=session,
#             #         seed=seed,  # 关联种子
#             #         idx=idx,
#             #         raw_text=f"[{level2.name}] 泛化用例 {idx+1} | seed={base} | prompt={effective_prompt[:30] if effective_prompt else 'default'}",
#             #     )
#             #     idx += 1
#             # 创建生成会话（不再保存 session.prompt；提示词只来自 level2.prompt）
#             session = GenerationSession.objects.create(
#                 level2=level2,
#                 prompt=None,
#                 model_name = getattr(settings, "ZHIPU_MODEL", "glm-4"),
#                 temperature=temperature,
#                 top_p=top_p,
#                 status="draft",
#                 created_by=request.user if request.user.is_authenticated else None
#             )
#
#             idx = 0
#             try:
#                 level1_name = level2.level1.name  # 注意：level2 需要能拿到 level1；如果没 select_related，也可以 level2.level1.name
#                 level2_name = level2.name
#                 scenario_prompt = level2.prompt or ""
#
#                 for config in seed_configs:
#                     seed_id = config.get("seed_id")
#                     n = int(config.get("n", 5) or 0)
#                     if n <= 0:
#                         continue
#
#                     try:
#                         seed = TestCaseSeed.objects.get(id=seed_id, level2=level2)
#                     except TestCaseSeed.DoesNotExist:
#                         continue
#
#                     GenerationSeedConfig.objects.create(session=session, seed=seed, n=n)
#
#                     # ✅ 真实调用大模型：一次 seed 生成 n 条
#                     cases = generate_cases_for_seed(
#                         level1_name=level1_name,
#                         level2_name=level2_name,
#                         seed_text=seed.text,
#                         prompt=scenario_prompt,  # 只有这一个提示词（来自二级功能）
#                         n=n,
#                         temperature=temperature,
#                         top_p=top_p,
#                         idx = idx,
#                     )
#
#                     for text in cases:
#                         GenerationItem.objects.create(
#                             session=session,
#                             seed=seed,
#                             idx=idx,
#                             raw_text=text,
#                         )
#                         idx += 1
#
#                 session.status = "done"
#                 session.save(update_fields=["status"])
#
#             except LLMError as e:
#                 session.status = "failed"
#                 session.save(update_fields=["status"])
#                 return JsonResponse({"error": str(e)}, status=500)
#             except Exception as e:
#                 session.status = "failed"
#                 session.save(update_fields=["status"])
#                 return JsonResponse({"error": f"生成失败: {str(e)}"}, status=500)
#
#     return JsonResponse({
#         "session_id": session.id,
#         "level2_id": level2.id,
#         "total": idx,
#         "message": f"生成完成！共生成 {idx} 条用例"
#     })

@require_http_methods(["POST"])
def workspace_generate(request):
    level2_id = request.POST.get("level2_id")
    seed_configs = request.POST.get("seed_configs")
    temperature = float(request.POST.get("temperature", 0.7))
    top_p = float(request.POST.get("top_p", 1.0))

    if not level2_id:
        return JsonResponse({"error": "缺少二级功能ID"}, status=400)
    if not seed_configs:
        return JsonResponse({"error": "请至少选择一个种子测试用例"}, status=400)

    import json
    try:
        seed_configs = json.loads(seed_configs)
        print("### workspace_generate: seed_configs_len =", len(seed_configs))
        print("### workspace_generate: seed_configs =", seed_configs[:5], "...")
    except Exception:
        return JsonResponse({"error": "种子配置格式错误"}, status=400)

    try:
        level2 = FeatureLevel2.objects.select_related("level1").get(id=level2_id)
    except FeatureLevel2.DoesNotExist:
        return JsonResponse({"error": "二级功能不存在"}, status=404)

    # ✅ 只用 level2.prompt 作为唯一 prompt
    scenario_prompt = level2.prompt or ""
    level1_name = level2.level1.name
    level2_name = level2.name

    with transaction.atomic():
        # ✅ 只创建一次 session
        session = GenerationSession.objects.create(
            level2=level2,
            prompt=None,
            model_name=getattr(settings, "ZHIPU_MODEL", "glm-4"),
            temperature=temperature,
            top_p=top_p,
            status="draft",
            created_by=request.user if request.user.is_authenticated else None
        )

        idx = 0
        try:
            for config in seed_configs:
                seed_id = config.get("seed_id")
                n = int(config.get("n", 0) or 0)
                if not seed_id or n <= 0:
                    continue

                try:
                    seed = TestCaseSeed.objects.get(id=seed_id, level2=level2)
                except TestCaseSeed.DoesNotExist:
                    continue

                # ✅ 不会触发唯一键冲突
                GenerationSeedConfig.objects.update_or_create(
                    session=session,
                    seed=seed,
                    defaults={"n": n},
                )

                # ✅ 只调用一次大模型
                cases = generate_cases_for_seed(
                    level1_name=level1_name,
                    level2_name=level2_name,
                    seed_text=seed.text,
                    prompt=scenario_prompt,
                    n=n,
                    temperature=temperature,
                    top_p=top_p,
                    idx=idx,
                )

                for text in cases:
                    GenerationItem.objects.create(
                        session=session,
                        seed=seed,
                        idx=idx,
                        raw_text=text,
                    )
                    idx += 1

            session.status = "done"
            session.save(update_fields=["status"])

        except LLMError as e:
            session.status = "failed"
            session.save(update_fields=["status"])
            return JsonResponse({"error": str(e)}, status=500)
        except Exception as e:
            session.status = "failed"
            session.save(update_fields=["status"])
            return JsonResponse({"error": f"生成失败: {str(e)}"}, status=500)

    return JsonResponse({
        "session_id": session.id,
        "level2_id": level2.id,
        "total": idx,
        "message": f"生成完成！共生成 {idx} 条用例"
    })

@require_http_methods(["POST"])
def delete_items(request):
    """批量删除并重排ID"""
    import json
    
    try:
        data = json.loads(request.body)
        level1_ids = data.get('level1_ids', [])
        level2_ids = data.get('level2_ids', [])
        seed_ids = data.get('seed_ids', [])
    except:
        return JsonResponse({"error": "请求数据格式错误"}, status=400)
    
    deleted_count = 0
    
    with transaction.atomic():
        # 1. 删除选中的一级功能（会级联删除二级功能和种子）
        if level1_ids:
            deleted_count += FeatureLevel1.objects.filter(id__in=level1_ids).count()
            FeatureLevel1.objects.filter(id__in=level1_ids).delete()
        
        # 2. 删除选中的二级功能（会级联删除种子）
        if level2_ids:
            deleted_count += FeatureLevel2.objects.filter(id__in=level2_ids).count()
            FeatureLevel2.objects.filter(id__in=level2_ids).delete()
        
        # 3. 删除选中的种子
        if seed_ids:
            deleted_count += TestCaseSeed.objects.filter(id__in=seed_ids).count()
            TestCaseSeed.objects.filter(id__in=seed_ids).delete()
        
        # 4. 重排ID
        if level1_ids:
            _reorder_ids(FeatureLevel1)
        if level2_ids:
            _reorder_ids(FeatureLevel2)
        if seed_ids:
            _reorder_ids(TestCaseSeed)
    
    return JsonResponse({
        "message": f"成功删除 {deleted_count} 项",
        "deleted_count": deleted_count
    })


def _reorder_ids(model):
    """重排模型的ID，使其连续"""
    from django.db import connection
    
    # 获取所有记录，按当前ID排序
    items = list(model.objects.all().order_by('id'))
    
    if not items:
        return
    
    # 临时禁用外键检查（MySQL）
    with connection.cursor() as cursor:
        cursor.execute('SET FOREIGN_KEY_CHECKS=0')
        
        try:
            # 重新分配ID
            for new_id, item in enumerate(items, start=1):
                if item.id != new_id:
                    # 更新ID
                    old_id = item.id
                    model.objects.filter(id=old_id).update(id=new_id)
            
            # 重置自增计数器
            table_name = model._meta.db_table
            next_id = len(items) + 1
            cursor.execute(f"ALTER TABLE {table_name} AUTO_INCREMENT = {next_id}")
        
        finally:
            # 重新启用外键检查
            cursor.execute('SET FOREIGN_KEY_CHECKS=1')


@require_http_methods(["POST"])
def update_level1(request):
    """更新一级功能"""
    import json
    
    try:
        data = json.loads(request.body)
        level1_id = data.get('id')
        name = data.get('name', '').strip()
        code = data.get('code', '').strip()
    except:
        return JsonResponse({"error": "请求数据格式错误"}, status=400)
    
    if not name:
        return JsonResponse({"error": "功能名称不能为空"}, status=400)
    
    try:
        level1 = FeatureLevel1.objects.get(id=level1_id)
        level1.name = name
        level1.code = code
        level1.save()
        return JsonResponse({"message": "更新成功", "name": level1.name})
    except FeatureLevel1.DoesNotExist:
        return JsonResponse({"error": "一级功能不存在"}, status=404)


@require_http_methods(["POST"])
def update_level2(request):
    """更新二级功能"""
    import json
    
    try:
        data = json.loads(request.body)
        level2_id = data.get('id')
        name = data.get('name', '').strip()
        prompt = data.get('prompt', '').strip()
    except:
        return JsonResponse({"error": "请求数据格式错误"}, status=400)
    
    if not name:
        return JsonResponse({"error": "功能名称不能为空"}, status=400)
    
    try:
        level2 = FeatureLevel2.objects.get(id=level2_id)
        level2.name = name
        level2.prompt = prompt if prompt else None
        level2.save()
        return JsonResponse({"message": "更新成功", "name": level2.name})
    except FeatureLevel2.DoesNotExist:
        return JsonResponse({"error": "二级功能不存在"}, status=404)


@require_http_methods(["POST"])
def update_seed(request):
    """更新种子测试用例"""
    import json
    
    try:
        data = json.loads(request.body)
        seed_id = data.get('id')
        text = data.get('text', '').strip()
    except:
        return JsonResponse({"error": "请求数据格式错误"}, status=400)
    
    if not text:
        return JsonResponse({"error": "测试用例内容不能为空"}, status=400)
    
    try:
        seed = TestCaseSeed.objects.get(id=seed_id)
        seed.text = text
        seed.save()
        return JsonResponse({"message": "更新成功"})
    except TestCaseSeed.DoesNotExist:
        return JsonResponse({"error": "种子用例不存在"}, status=404)


@require_http_methods(["POST"])
@require_http_methods(["POST"])
def regenerate_item(request):
    """
    重新生成单条测试用例
    ⭐ 关键修改：创建新记录并关联到原始记录
    """
    import json

    try:
        data = json.loads(request.body)
        item_id = data.get('item_id')
    except:
        return JsonResponse({"error": "请求数据格式错误"}, status=400)

    try:
        # 获取原始item
        original_item = GenerationItem.objects.get(id=item_id)

        # # Mock生成新文本（实际应该调用AI模型）
        # import random
        # seed_text = original_item.seed.text if original_item.seed else "无种子"
        # new_text = f"{seed_text} [重新生成-{random.randint(100, 999)}]"
        try:
            session = original_item.session
            level2 = session.level2
            level1_name = level2.level1.name
            level2_name = level2.name
            scenario_prompt = level2.prompt or ""
            seed_text = original_item.seed.text if original_item.seed else ""

            if not seed_text:
                return JsonResponse({"error": "该条记录没有关联种子，无法重新生成"}, status=400)

            # ✅ 真实调用：生成 1 条
            new_text = generate_cases_for_seed(
                level1_name=level1_name,
                level2_name=level2_name,
                seed_text=seed_text,
                prompt=scenario_prompt,
                n=1,
                temperature=session.temperature,
                top_p=session.top_p,
                idx='重试生成'

            )[0]
        except LLMError as e:
            return JsonResponse({"error": str(e)}, status=500)

        # ⭐ 创建新记录（重新生成模式）
        new_item = GenerationItem.objects.create(
            session=original_item.session,
            seed=original_item.seed,
            idx=original_item.idx,  # 保持相同的idx，表示这是同一个位置的重新生成
            raw_text=new_text,
            edited_text=None,  # 重新生成后清空编辑内容
            is_edited=False,
            regen_from_item=original_item  # 关联到原始记录
        )

        return JsonResponse({
            "new_text": new_text,
            "item_id": new_item.id,  # 返回新创建的记录ID
            "message": "重新生成成功"
        })
    except GenerationItem.DoesNotExist:
        return JsonResponse({"error": "生成项不存在"}, status=404)
    except Exception as e:
        return JsonResponse({"error": f"重新生成失败: {str(e)}"}, status=500)


@require_http_methods(["POST"])
def save_all_edits(request):
    """批量保存所有编辑"""
    import json
    
    try:
        data = json.loads(request.body)
        session_id = data.get('session_id')
        updates = data.get('updates', [])
    except:
        return JsonResponse({"error": "请求数据格式错误"}, status=400)
    
    if not updates:
        return JsonResponse({"error": "没有需要保存的修改"}, status=400)
    
    try:
        session = GenerationSession.objects.get(id=session_id)
    except GenerationSession.DoesNotExist:
        return JsonResponse({"error": "会话不存在"}, status=404)
    
    saved_count = 0
    
    with transaction.atomic():
        for update in updates:
            item_id = update.get('item_id')
            text = update.get('text', '').strip()
            
            if not text:
                continue
            
            try:
                item = GenerationItem.objects.get(id=item_id, session=session)
                item.edited_text = text
                item.is_edited = True
                item.save()
                saved_count += 1
            except GenerationItem.DoesNotExist:
                continue
    
    return JsonResponse({
        "message": f"成功保存 {saved_count} 条修改",
        "saved_count": saved_count
    })


@require_http_methods(["POST"])
def save_to_final(request):
    """
    将生成结果保存到最终库
    ⭐ 关键修改：按idx分组，找到每个idx的最新版本
    功能流程：
    1. 获取该会话的所有GenerationItem记录
    2. 按idx分组，对每个idx找到最新的GenerationItem记录（按created_at倒序）
    3. 使用该记录的final_text作为最终用例
    4. 保存到SavedCaseItem表
    """
    import json
    from datetime import datetime
    from collections import defaultdict

    try:
        data = json.loads(request.body)
        session_id = data.get('session_id')
        version_title = data.get('title', '')  # 版本名称
        status = data.get('status', 'confirmed')  # draft/confirmed/delivered
    except:
        return JsonResponse({"error": "请求数据格式错误"}, status=400)

    try:
        session = GenerationSession.objects.get(id=session_id)

        # 生成唯一的批次ID: level2_{id}_{timestamp}
        timestamp = datetime.now().strftime('%Y%m%d%H%M%S')
        saved_batch_id = f"level2_{session.level2_id}_{timestamp}"

        # 生成默认标题
        if not version_title:
            version_title = f"{session.level2.name} - {datetime.now().strftime('%Y-%m-%d %H:%M:%S')}"

        with transaction.atomic():
            # 获取该场景的所有旧记录数量（用于提示）
            old_count = SavedCaseItem.objects.filter(level2=session.level2).count()

            # ⭐ 关键修改：按idx分组，找到每个idx的最新GenerationItem
            all_items = GenerationItem.objects.filter(session=session).order_by('idx', '-created_at')

            if not all_items.exists():
                return JsonResponse({"error": "该会话没有生成任何用例"}, status=400)

            # 按idx分组
            items_by_idx = defaultdict(list)
            for item in all_items:
                items_by_idx[item.idx].append(item)

            saved_count = 0
            final_idx = 1  # SavedCaseItem的索引，从1开始

            # 按idx排序后遍历
            for idx in sorted(items_by_idx.keys()):
                items_list = items_by_idx[idx]
                # 取该idx的最新记录（第一个，因为已经按created_at倒序排序）
                latest_item = items_list[0]

                # 使用该记录的final_text
                final_text = latest_item.final_text

                # 保存到最终库
                SavedCaseItem.objects.create(
                    level2=session.level2,
                    from_session=session,
                    from_gen_item=latest_item,
                    saved_batch_id=saved_batch_id,
                    idx=final_idx,
                    text=final_text,
                    version_title=version_title,
                    status=status,
                    created_by=request.user if request.user.is_authenticated else None
                )

                saved_count += 1
                final_idx += 1

            if saved_count == 0:
                return JsonResponse({"error": "没有找到可保存的用例"}, status=400)

        message = f"成功保存 {saved_count} 条测试用例到最终库"
        if old_count > 0:
            message += f"（该场景原有 {old_count} 条记录）"

        return JsonResponse({
            "success": True,
            "saved_batch_id": saved_batch_id,
            "saved_count": saved_count,
            "old_count": old_count,
            "version_title": version_title,
            "message": message
        })
    except GenerationSession.DoesNotExist:
        return JsonResponse({"error": "会话不存在"}, status=404)
    except Exception as e:
        return JsonResponse({"error": f"保存到最终库失败: {str(e)}"}, status=500)


# @require_http_methods(["GET"])
# def get_level2_list(request):
#     """AJAX接口：根据一级功能ID获取二级功能列表"""
#     level1_id = request.GET.get("level1_id")
#     if not level1_id:
#         return JsonResponse({"error": "缺少level1_id参数"}, status=400)
#
#     try:
#         level1 = FeatureLevel1.objects.get(id=level1_id)
#         level2_list = FeatureLevel2.objects.filter(level1=level1).order_by("name")
#         data = [{"id": l2.id, "name": l2.name} for l2 in level2_list]
#         return JsonResponse({"level2_list": data})
#     except FeatureLevel1.DoesNotExist:
#         return JsonResponse({"error": "一级功能不存在"}, status=404)


def level2_detail(request, level2_id):
    """
    二级功能详情：展示生成结果（新版美化界面）
    1. 只读展示种子测试用例
    2. 按种子分组展示所有生成结果（包括重新生成的版本）
    """
    level2 = get_object_or_404(FeatureLevel2.objects.select_related("level1"), id=level2_id)
    seeds = TestCaseSeed.objects.filter(level2=level2).order_by("created_at")

    # 取最新 session
    session = GenerationSession.objects.filter(level2=level2).order_by("-created_at").first()

    items_by_seed = {}
    total_count = 0

    if session:
        # 获取该会话的所有生成项，按种子分组显示所有记录
        items = GenerationItem.objects.filter(session=session).select_related('seed').order_by('seed_id', 'created_at')

        for item in items:
            if item.seed:
                if item.seed not in items_by_seed:
                    items_by_seed[item.seed] = []
                items_by_seed[item.seed].append(item)
                total_count += 1

    context = {
        "level2": level2,
        "seeds": seeds,
        "session": session,
        "items_by_seed": items_by_seed,
        "total_count": total_count,
    }
    return render(request, "generation_result.html", context)


def session_edit_and_save(request, session_id):
    """
    编辑 session 的 items（5条）并保存到 SavedCaseSet + SavedCaseItem
    """
    session = get_object_or_404(GenerationSession.objects.select_related("level2", "level2__level1"), id=session_id)
    items_qs = GenerationItem.objects.filter(session=session).order_by("idx")

    if request.method != "POST":
        return redirect(reverse("Generate_testcases:level2_detail", args=[session.level2_id]))

    action = request.POST.get("action")

    # 编辑：保存编辑内容（不创建 SavedCaseSet）
    if action == "save_edits":
        formset = GenerationItemFormSet(request.POST, queryset=items_qs)
        if formset.is_valid():
            with transaction.atomic():
                objs = formset.save(commit=False)
                # 将 is_edited 置为 True（只要有 edited_text）
                for obj in objs:
                    if obj.edited_text is not None:
                        obj.is_edited = True
                    obj.save()
            messages.success(request, "编辑已保存")
        else:
            messages.error(request, "编辑保存失败，请检查输入")
        return redirect(reverse("Generate_testcases:level2_detail", args=[session.level2_id]))

    # 保存：把最终 5 条写入新表 SavedCaseSet/SavedCaseItem
    if action == "final_save":
        formset = GenerationItemFormSet(request.POST, queryset=items_qs)
        save_form = SaveCaseSetForm(request.POST)

        if formset.is_valid() and save_form.is_valid():
            with transaction.atomic():
                # 先保存编辑内容
                objs = formset.save(commit=False)
                for obj in objs:
                    if obj.edited_text is not None:
                        obj.is_edited = True
                    obj.save()

                case_set = save_form.save(commit=False)
                case_set.level2 = session.level2
                case_set.from_session = session
                case_set.created_by = request.user if request.user.is_authenticated else None
                case_set.save()

                # 写入 5 条最终文本
                items = GenerationItem.objects.filter(session=session).order_by("idx")
                for it in items:
                    SavedCaseItem.objects.create(
                        case_set=case_set,
                        idx=it.idx,
                        text=it.final_text,
                        from_gen_item=it,
                    )

            messages.success(request, "已保存到新表（SavedCaseSet/SavedCaseItem）")
        else:
            messages.error(request, "保存失败，请检查输入")
        return redirect(reverse("Generate_testcases:level2_detail", args=[session.level2_id]))

    # 未知 action
    return redirect(reverse("Generate_testcases:level2_detail", args=[session.level2_id]))





@require_http_methods(["POST"])
def import_excel_to_db(request):
    """
    Excel 导入规则（固定列位）：
    - 忽略第 1 列
    - 第 2 列：一级功能
    - 第 3 列：二级功能
    - 第 4 列：二级功能场景提示词 prompt（支持合并单元格向下继承）
    - 第 5 列：种子测试用例 seed
    说明：prompt 可以为空；seed 不继承
    """
    print("### import_excel_to_db FINAL VERSION (v3-safe) ###")

    try:
        f = request.FILES.get("file")
        if not f:
            return JsonResponse({"ok": False, "msg": "未收到文件 file"}, status=400)

        if not f.name.lower().endswith(".xlsx"):
            return JsonResponse({"ok": False, "msg": "目前只支持 .xlsx 格式"}, status=400)

        try:
            wb = load_workbook(f, data_only=True)
            ws = wb.active
        except Exception as e:
            return JsonResponse({"ok": False, "msg": f"读取Excel失败：{e}"}, status=400)

        def norm(v):
            return "" if v is None else str(v).strip()

        rows = list(ws.iter_rows(values_only=True))
        if len(rows) <= 1:
            return JsonResponse({"ok": False, "msg": "Excel中没有数据行"}, status=400)

        created_l1 = created_l2 = created_seed = updated_l2_prompt = skipped = 0

        # 合并单元格继承：l1/l2/prompt
        last_l1 = ""
        last_l2 = ""
        last_prompt = ""

        with transaction.atomic():
            for r in rows[1:]:
                # B/C/D/E
                raw_l1 = norm(r[1]) if len(r) > 1 else ""
                raw_l2 = norm(r[2]) if len(r) > 2 else ""
                raw_prompt = norm(r[3]) if len(r) > 3 else ""   # 第4列 prompt
                seed = norm(r[4]) if len(r) > 4 else ""         # 第5列 seed

                # l1/l2 向下继承
                l1 = raw_l1 or last_l1
                l2 = raw_l2 or last_l2

                # 必须有 l1/l2，prompt 可为空
                if not l1 or not l2:
                    skipped += 1
                    continue

                # prompt 向下继承（只对合并单元格有效：同一个 l1+l2 的连续行）
                if raw_prompt:
                    prompt = raw_prompt
                else:
                    prompt = last_prompt if (l1 == last_l1 and l2 == last_l2) else ""

                # 更新缓存
                last_l1, last_l2, last_prompt = l1, l2, prompt

                level1, l1_created = FeatureLevel1.objects.get_or_create(name=l1)
                if l1_created:
                    created_l1 += 1

                level2, l2_created = FeatureLevel2.objects.get_or_create(level1=level1, name=l2)
                if l2_created:
                    created_l2 += 1

                # prompt 可为空：为空不覆盖；有值才更新
                if prompt and (level2.prompt or "") != prompt:
                    level2.prompt = prompt
                    level2.save(update_fields=["prompt"])
                    updated_l2_prompt += 1

                # seed 写入（不继承；为空则跳过）
                if seed:
                    _, seed_created = TestCaseSeed.objects.get_or_create(level2=level2, text=seed)
                    if seed_created:
                        created_seed += 1

        return JsonResponse({
            "ok": True,
            "msg": "导入成功",
            "stats": {
                "created_level1": created_l1,
                "created_level2": created_l2,
                "updated_level2_prompt": updated_l2_prompt,
                "created_seed": created_seed,
                "skipped_rows": skipped,
            }
        })

    except Exception as e:
        # ✅ 兜底：保证永远返回 HttpResponse，不会是 None
        return JsonResponse({"ok": False, "msg": f"导入异常：{str(e)}"}, status=500)


